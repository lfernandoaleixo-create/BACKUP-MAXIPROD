import openpyxl

wb = openpyxl.load_workbook('/home/ubuntu/grupo-fox-dashboard/scripts/flor_minas.xlsx', data_only=False)
ws = wb['COTAÇÃO FLOR DE MINAS ']

# Extract all cities (rows 55 to 166)
print("=== TABELA DE CIDADES E PRAZOS (rows 55-166) ===")
cities = []
for row_idx in range(55, 167):
    city = ws.cell(row=row_idx, column=1).value
    state = ws.cell(row=row_idx, column=2).value
    prazo = ws.cell(row=row_idx, column=3).value
    if city:
        cities.append({'cidade': city, 'estado': state, 'prazo': prazo})
        print(f"  {city} - {state}: {prazo}")

print(f"\nTotal cities: {len(cities)}")

# Extract pricing table (rows 47-52)
print("\n=== TABELA DE PREÇOS (rows 47-52) ===")
for row_idx in range(47, 53):
    row_data = []
    for col_idx in range(1, 10):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_data.append(f"{col_letter}={repr(val)}")
    if row_data:
        print(f"  Row {row_idx}: {', '.join(row_data)}")

# Extract input fields
print("\n=== CAMPOS DE ENTRADA ===")
print(f"  C11 (label): {ws.cell(row=11, column=3).value}")
print(f"  C12 (cidade selecionada): {ws.cell(row=12, column=3).value}")
print(f"  C14 (label): {ws.cell(row=14, column=3).value}")
print(f"  G14 (label): {ws.cell(row=14, column=7).value}")
print(f"  C15 (valor NF): {ws.cell(row=15, column=3).value}")
print(f"  G15 (peso kg): {ws.cell(row=15, column=7).value}")

# Extract output fields
print("\n=== CAMPOS DE SAÍDA ===")
print(f"  C19 (label): {ws.cell(row=19, column=3).value}")
print(f"  G19 (label): {ws.cell(row=19, column=7).value}")
print(f"  C20 (valor frete - formula): {ws.cell(row=20, column=3).value}")
print(f"  F20 (% do frete - formula): {ws.cell(row=20, column=6).value}")
print(f"  G20 (prazo - formula): {ws.cell(row=20, column=7).value}")

# Check H47 value
print(f"\n=== H47 (volumes): {ws.cell(row=47, column=8).value}")

# Also check if there's data beyond row 166
print("\n=== DATA BEYOND ROW 166 ===")
for row_idx in range(167, 179):
    row_data = []
    for col_idx in range(1, 10):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_data.append(f"{col_letter}={repr(val)}")
    if row_data:
        print(f"  Row {row_idx}: {', '.join(row_data)}")

# Also load with data_only=True to see computed values
wb2 = openpyxl.load_workbook('/home/ubuntu/grupo-fox-dashboard/scripts/flor_minas.xlsx', data_only=True)
ws2 = wb2['COTAÇÃO FLOR DE MINAS ']
print("\n=== COMPUTED VALUES (cached) ===")
print(f"  C20 (valor frete): {ws2.cell(row=20, column=3).value}")
print(f"  F20 (% frete): {ws2.cell(row=20, column=6).value}")
print(f"  G20 (prazo): {ws2.cell(row=20, column=7).value}")
print(f"  H48: {ws2.cell(row=48, column=8).value}")
print(f"  I48: {ws2.cell(row=48, column=9).value}")
print(f"  I49: {ws2.cell(row=49, column=9).value}")
print(f"  I50: {ws2.cell(row=50, column=9).value}")
print(f"  I51: {ws2.cell(row=51, column=9).value}")
print(f"  I52 (total): {ws2.cell(row=52, column=9).value}")
print(f"  C51 (peso*0.747): {ws2.cell(row=51, column=3).value}")
