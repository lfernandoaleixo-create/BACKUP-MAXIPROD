import openpyxl
import json

wb = openpyxl.load_workbook('/home/ubuntu/grupo-fox-dashboard/scripts/flor_minas.xlsx', data_only=False)
print('Sheets:', wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n{"="*60}')
    print(f'Sheet: {name}')
    print(f'Dimensions: {ws.dimensions}')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    print(f'{"="*60}')
    
    # Print all rows with data (up to 100 rows)
    print('\nAll data (up to 100 rows):')
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=False), 1):
        cells_with_data = []
        for cell in row:
            if cell.value is not None:
                cells_with_data.append(f'{cell.coordinate}={repr(cell.value)}')
        if cells_with_data:
            print(f'  Row {row_idx}: {", ".join(cells_with_data)}')
    
    # Print formulas separately
    print('\nFormulas found:')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                print(f'  {cell.coordinate}: {cell.value}')
    
    # Print merged cells
    if ws.merged_cells.ranges:
        print(f'\nMerged cells: {list(ws.merged_cells.ranges)}')
